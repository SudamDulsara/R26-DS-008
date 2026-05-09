"""
io/writer.py
------------
Writes processed Documents to an Excel file for human review.

Two sheets:
  - "Documents": one row per processed doc, with raw vs cleaned text,
    dedup info, verdict, and reasons. Verdicts are colour-coded.
  - "Summary":   counts and aggregates for quick at-a-glance review.

This is a *reporting* output, not the production handoff. The real corpus
output (for training) will be a JSONL/Parquet writer added later.

Requires: openpyxl  (pip install openpyxl)
"""

from collections import Counter
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..schema import Document, Verdict


_VERDICT_FILL = {
    Verdict.ACCEPT:  PatternFill("solid", fgColor="C6EFCE"),  # green
    Verdict.REJECT:  PatternFill("solid", fgColor="FFC7CE"),  # red
    Verdict.REVIEW:  PatternFill("solid", fgColor="FFEB9C"),  # yellow
    Verdict.PENDING: PatternFill("solid", fgColor="D9D9D9"),  # grey
}

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="305496")


class XlsxWriter:
    """Write a batch of processed Documents to an .xlsx file."""

    COLUMNS = [
        ("doc_id",            12),
        ("source",            10),
        ("verdict",           10),
        ("raw_text",          50),
        ("cleaned_text",      50),
        ("chars_removed",     14),
        ("content_hash",      18),
        ("is_duplicate",      14),
        ("verdict_reasons",   30),
        ("stages_completed",  35),
    ]

    def __init__(self, path: str | Path):
        self.path = Path(path)

    def write(self, docs: Iterable[Document]) -> None:
        docs = list(docs)
        wb = Workbook()
        self._write_documents_sheet(wb.active, docs)
        self._write_summary_sheet(wb.create_sheet("Summary"), docs)
        wb.save(self.path)

    def _write_documents_sheet(self, ws, docs: list[Document]) -> None:
        ws.title = "Documents"

        for col_idx, (name, width) in enumerate(self.COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = "A2"

        for row_idx, doc in enumerate(docs, start=2):
            row_values = [
                doc.doc_id,
                doc.source.value,
                doc.verdict.value,
                doc.raw_text,
                doc.text,
                doc.linguistic.get("chars_removed", ""),
                (doc.quality.get("content_hash", "") or "")[:16],
                "yes" if doc.quality.get("is_duplicate") else "no",
                "; ".join(doc.verdict_reasons),
                " → ".join(doc.stages_completed),
            ]
            for col_idx, value in enumerate(row_values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

            verdict_cell = ws.cell(row=row_idx, column=3)
            verdict_cell.fill = _VERDICT_FILL.get(doc.verdict, _VERDICT_FILL[Verdict.PENDING])
            verdict_cell.font = Font(bold=True)
            verdict_cell.alignment = Alignment(horizontal="center", vertical="center")

    def _write_summary_sheet(self, ws, docs: list[Document]) -> None:
        verdict_counts = Counter(d.verdict.value for d in docs)
        source_counts = Counter(d.source.value for d in docs)
        total_chars_removed = sum(d.linguistic.get("chars_removed", 0) for d in docs)
        duplicates = sum(1 for d in docs if d.quality.get("is_duplicate"))

        rows = [
            ("Sinhala Quality Pipeline — Run Summary", ""),
            ("", ""),
            ("Total documents processed", len(docs)),
            ("Total characters removed by normalization", total_chars_removed),
            ("Exact duplicates detected", duplicates),
            ("", ""),
            ("By verdict:", ""),
            *[(f"  {v}", n) for v, n in sorted(verdict_counts.items())],
            ("", ""),
            ("By source:", ""),
            *[(f"  {s}", n) for s, n in sorted(source_counts.items())],
        ]

        for row_idx, (label, value) in enumerate(rows, start=1):
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=value)
            if row_idx == 1:
                ws.cell(row=row_idx, column=1).font = Font(bold=True, size=14)

        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 15